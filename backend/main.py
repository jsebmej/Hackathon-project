"""API y aplicación web local de DataLex 1581.

Este módulo integra el motor que antes vivía en ``datalex_auditoria.py``:
generación de criterios, entrevista adaptativa, clasificación, puntuación,
persistencia tolerante a fallos y exportación a Excel/JSON.
"""

from __future__ import annotations

import asyncio
import base64
import datetime as dt
import hashlib
import hmac
import json
import logging
import os
import re
import secrets
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from typing import Annotated, Any

import httpx
from dotenv import load_dotenv
from fastapi import Depends, FastAPI, HTTPException, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from fastapi.staticfiles import StaticFiles
from google.auth.transport import requests as google_requests
from google.oauth2 import id_token
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel, Field

try:
    import psycopg
except ImportError:  # La aplicación y el Excel siguen funcionando sin PostgreSQL.
    psycopg = None


BASE_DIR = Path(__file__).resolve().parent
PROJECT_DIR = BASE_DIR.parent
FRONTEND_DIR = PROJECT_DIR / "frontend"
load_dotenv(BASE_DIR / ".env")

GOOGLE_CLIENT_ID = os.getenv("GOOGLE_CLIENT_ID", "").strip()
LOCAL_AUTH_USERNAME = os.getenv("LOCAL_AUTH_USERNAME", "").strip()
LOCAL_AUTH_PASSWORD = os.getenv("LOCAL_AUTH_PASSWORD", "")
SESSION_TTL_HOURS = int(os.getenv("SESSION_TTL_HOURS", "12"))
OLLAMA_BASE_URL = os.getenv("OLLAMA_BASE_URL", "http://localhost:11434").rstrip("/")
OLLAMA_MODEL = os.getenv("OLLAMA_MODEL", "datalex1581")
AUDIT_DIR_CONFIG = Path(os.getenv("AUDIT_LOG_DIR", "auditorias"))
AUDIT_DIR = AUDIT_DIR_CONFIG if AUDIT_DIR_CONFIG.is_absolute() else BASE_DIR / AUDIT_DIR_CONFIG
FRONTEND_ORIGINS = [
    origin.strip()
    for origin in os.getenv(
        "FRONTEND_ORIGINS",
        "http://localhost:5173,http://127.0.0.1:5173",
    ).split(",")
    if origin.strip()
]

logger = logging.getLogger("datalex")

app = FastAPI(title="DataLex 1581", version="1.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=FRONTEND_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["Content-Disposition"],
)
bearer_scheme = HTTPBearer(auto_error=False)


# ---------------------------------------------------------------------------
# Motor de auditoría (integrado desde datalex_auditoria.py)
# ---------------------------------------------------------------------------
DOMINIOS = [
    "Consentimiento",
    "Finalidad",
    "Seguridad",
    "Derechos del titular",
    "Gestión documental",
]
CRIT_MIN, CRIT_MAX = 4, 6
TOPE_POR_DOMINIO = 10
PESO = {"Alta": 3, "Media": 2, "Baja": 1}
VALOR = {"Sí": 1.0, "Parcial": 0.5, "No": 0.0}
UMBRAL_ALTO, UMBRAL_MEDIO = 0.80, 0.50
ARCHIVO_SEMILLA = "20260626_-_Criterios_de_aceptacion.xlsx"

SEMILLA = {
    "Consentimiento": [
        "La empresa solicita autorización previa del titular",
        "La autorización es informada y clara",
        "Se conserva evidencia del consentimiento",
        "Permite revocar el consentimiento",
        "El consentimiento es verificable digital o físicamente",
        "No trata datos sin autorización válida",
        "Existen formatos de autorización estandarizados",
        "Se valida el consentimiento antes del tratamiento",
        "Se actualizan consentimientos cuando cambia la finalidad",
        "Se informa al titular sobre el uso de datos",
    ],
    "Finalidad": [
        "Los datos tienen una finalidad claramente definida",
        "La finalidad se informa al titular antes del tratamiento",
        "No se usan datos para fines distintos a los autorizados",
        "La finalidad está documentada en políticas internas",
        "Se revisa periódicamente la finalidad del tratamiento",
        "Los sistemas validan el uso según la finalidad",
        "Se eliminan datos cuando se cumple la finalidad",
        "Los empleados conocen la finalidad del tratamiento",
        "Se audita el cumplimiento de la finalidad",
        "Se actualizan finalidades cuando cambian procesos",
    ],
    "Seguridad": [
        "Existen controles de acceso a bases de datos",
        "Se usa cifrado de información sensible",
        "Hay copias de seguridad periódicas",
        "Existen políticas de seguridad de la información",
        "Se monitorean accesos no autorizados",
        "Se gestionan incidentes de seguridad",
        "Se restringe acceso según roles",
        "Se protegen credenciales de acceso",
        "Se realizan pruebas de seguridad periódicas",
        "Los sistemas están protegidos contra ataques externos",
    ],
    "Derechos del titular": [
        "La empresa responde solicitudes de consulta",
        "La empresa responde solicitudes de reclamo",
        "Permite actualización de datos personales",
        "Permite eliminación de datos",
        "Existen canales de atención al titular",
        "Se respetan tiempos legales de respuesta",
        "Se registra cada solicitud del titular",
        "Se verifica identidad del solicitante",
        "Se informa al titular sobre el uso de sus datos",
        "Existe procedimiento documentado de derechos",
    ],
    "Gestión documental": [
        "Existe política de tratamiento de datos",
        "La política está publicada y accesible",
        "Se actualiza la política periódicamente",
        "Se documentan bases de datos",
        "Existe inventario de datos personales",
        "Se documentan responsables del tratamiento",
        "Se documentan encargados externos",
        "Se realizan auditorías internas",
        "Se conservan evidencias de cumplimiento",
        "Existe control de versiones de documentos",
    ],
}

PALABRAS_DUDA = (
    "no entiendo",
    "no se",
    "no sé",
    "que es",
    "qué es",
    "explica",
    "explícame",
    "aclara",
    "ayuda",
    "no comprendo",
    "como asi",
    "cómo así",
    "a que te refieres",
    "a qué te refieres",
    "no me queda claro",
    "ejemplo",
    "puedes explicar",
)


def norm(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value)).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]", "", text.lower())


def normaliza_valor(value: Any) -> str | None:
    normalized = norm(value)
    if normalized == "si":
        return "Sí"
    if normalized.startswith("par"):
        return "Parcial"
    if normalized == "no":
        return "No"
    return None


def normaliza_importancia(value: Any) -> str:
    normalized = norm(value)
    if normalized.startswith("alt"):
        return "Alta"
    if normalized.startswith("baj"):
        return "Baja"
    return "Media"


def _limpia(texto: str) -> str:
    return re.sub(r"```[a-zA-Z]*", "", texto).replace("```", "")


def extraer_array(texto: str) -> list[Any] | None:
    limpio = _limpia(texto)
    inicio, fin = limpio.find("["), limpio.rfind("]")
    if inicio == -1 or fin == -1:
        return None
    fragmento = re.sub(r",(\s*[\]}])", r"\1", limpio[inicio : fin + 1])
    try:
        value = json.loads(fragmento)
    except json.JSONDecodeError:
        return None
    return value if isinstance(value, list) else None


def extraer_objeto(texto: str) -> dict[str, Any] | None:
    limpio = _limpia(texto)
    inicio, fin = limpio.find("{"), limpio.rfind("}")
    if inicio == -1 or fin == -1:
        return None
    fragmento = re.sub(r",(\s*[}\]])", r"\1", limpio[inicio : fin + 1])
    try:
        value = json.loads(fragmento)
    except json.JSONDecodeError:
        return None
    return value if isinstance(value, dict) else None


def cargar_semilla() -> dict[str, list[str]]:
    """Lee una matriz externa cuando existe; si falla, usa la semilla embebida."""
    candidates = [BASE_DIR / ARCHIVO_SEMILLA, PROJECT_DIR / ARCHIVO_SEMILLA]
    ruta = next((candidate for candidate in candidates if candidate.exists()), None)
    if ruta is None:
        return SEMILLA
    try:
        worksheet = load_workbook(ruta, read_only=True, data_only=True).active
        base: dict[str, list[str]] = {}
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            dominio = row[0] if row else None
            criterio = row[1] if len(row) > 1 else None
            if dominio and criterio:
                base.setdefault(str(dominio).strip(), []).append(str(criterio).strip())
        return base or SEMILLA
    except Exception as exc:
        logger.warning("No fue posible cargar la semilla externa: %s", exc)
        return SEMILLA


class OllamaError(RuntimeError):
    pass


async def pedir(instruccion: str, timeout: float = 300.0) -> str:
    """Ejecuta una micro-tarea contra el modelo local configurado en Ollama."""
    try:
        async with httpx.AsyncClient(timeout=timeout) as client:
            response = await client.post(
                f"{OLLAMA_BASE_URL}/api/chat",
                json={
                    "model": OLLAMA_MODEL,
                    "messages": [{"role": "user", "content": instruccion}],
                    "stream": False,
                    "options": {"temperature": 0.2},
                },
            )
            response.raise_for_status()
    except httpx.HTTPError as exc:
        raise OllamaError(f"No fue posible consultar Ollama: {exc}") from exc

    contenido = response.json().get("message", {}).get("content", "").strip()
    if not contenido:
        raise OllamaError("Ollama respondió sin contenido")
    return contenido


def perfil_a_texto(profile: "AuditProfile") -> str:
    yes_no = lambda value: "Sí" if value else "No"
    return (
        f"Empresa: {profile.empresa}. NIT: {profile.nit or 'No informado'}. "
        f"Sector: {profile.sector}. Tamaño: {profile.tamano_empresa}. "
        f"Datos sensibles: {yes_no(profile.datos_sensibles)}. "
        f"Datos de menores: {yes_no(profile.datos_menores)}. "
        f"Transferencias internacionales: {yes_no(profile.transferencias_internacionales)}. "
        f"Canales digitales: {yes_no(profile.canales_digitales)}. "
        f"Notas: {profile.notas or 'ninguna'}."
    )


async def generar_criterios(
    perfil: str, dominio: str, semilla_dominio: list[str]
) -> list[dict[str, Any]]:
    base = "\n".join(f"- {criterio}" for criterio in semilla_dominio)
    instruccion = (
        f"Perfil de la empresa:\n{perfil}\n\n"
        f'Dominio: "{dominio}" (Ley 1581 de 2012).\n'
        f"Criterios base de auditoría para este dominio:\n{base}\n\n"
        "Selecciona los criterios que apliquen a esta empresa, hazlos específicos y agrega "
        "solo los que resulten necesarios por su sector y riesgos. No inventes artículos ni sanciones. "
        f"Entrega entre {CRIT_MIN} y {CRIT_MAX} criterios y asigna una importancia. "
        'Responde SOLO JSON: [{"criterio":"...","importancia":"Alta|Media|Baja"}]'
    )
    salida: list[dict[str, Any]] = []
    try:
        items = extraer_array(await pedir(instruccion)) or []
        for item in items:
            if not isinstance(item, dict):
                continue
            criterio = next((v for k, v in item.items() if norm(k) == "criterio"), None)
            importancia = next((v for k, v in item.items() if norm(k) == "importancia"), None)
            if criterio and str(criterio).strip():
                imp = normaliza_importancia(importancia)
                salida.append(
                    {
                        "dominio": dominio,
                        "criterio": str(criterio).strip(),
                        "importancia": imp,
                        "peso": PESO[imp],
                    }
                )
    except Exception as exc:
        logger.warning("Se usará la semilla para %s porque Ollama falló: %s", dominio, exc)

    if not salida:
        salida = [
            {
                "dominio": dominio,
                "criterio": criterio,
                "importancia": "Media",
                "peso": PESO["Media"],
            }
            for criterio in semilla_dominio[:CRIT_MAX]
        ]
    return salida[:TOPE_POR_DOMINIO]


async def redactar_pregunta(criterio: str) -> str:
    respaldo = f"Respecto a «{criterio}», ¿cuál es la situación actual en la empresa?"
    try:
        respuesta = await pedir(
            "Escribe una sola pregunta clara y natural en español para verificar en una empresa "
            f"este criterio: \"{criterio}\". Devuelve únicamente la pregunta."
        )
        return respuesta.splitlines()[0].strip() or respaldo
    except Exception as exc:
        logger.warning("No fue posible redactar una pregunta con Ollama: %s", exc)
        return respaldo


async def clasificar(criterio: str, respuesta: str) -> dict[str, str] | None:
    instruccion = (
        f'Criterio: "{criterio}".\nRespuesta del auditado: "{respuesta}".\n\n'
        "Clasifica tres dimensiones y responde SOLO este JSON:\n"
        '{"Cumple":"Sí|No|Parcial","Documentacion":"Sí|No|Parcial",'
        '"Controles":"Sí|No|Parcial"}\n'
        "Cumple indica si satisface el criterio; Documentacion si existe soporte documental; "
        "Controles si existen medidas que lo aseguren. Si algo no se menciona, no lo asumas como Sí."
    )
    try:
        obj = extraer_objeto(await pedir(instruccion))
    except Exception as exc:
        logger.warning("Clasificación con Ollama no disponible: %s", exc)
        return None
    if not obj:
        return None

    resultado: dict[str, str] = {}
    for source, destination in (
        ("cumple", "Cumple"),
        ("documentacion", "Documentacion"),
        ("controles", "Controles"),
    ):
        value = next((v for k, v in obj.items() if norm(k) == source), None)
        normalized = normaliza_valor(value)
        if normalized is None:
            return None
        resultado[destination] = normalized
    return resultado


def clasificar_respaldo(respuesta: str) -> dict[str, str]:
    """Mantiene la evaluación activa cuando Ollama no logra devolver JSON válido."""
    text = unicodedata.normalize("NFKD", respuesta.lower()).encode("ascii", "ignore").decode()
    negative = bool(re.search(r"\b(no|nunca|ningun|ninguna|carece|inexistente)\b", text))
    partial = any(word in text for word in ("parcial", "proceso", "algunos", "informal", "a veces"))
    if negative and not partial:
        cumple = "No"
    elif partial:
        cumple = "Parcial"
    else:
        cumple = "Sí"

    documentation_words = ("document", "politica", "manual", "registro", "evidencia", "formato")
    control_words = ("control", "revision", "auditor", "monitore", "valid", "responsable", "acceso")
    documentacion = cumple if any(word in text for word in documentation_words) else (
        "No" if cumple == "No" else "Parcial"
    )
    controles = cumple if any(word in text for word in control_words) else (
        "No" if cumple == "No" else "Parcial"
    )
    return {"Cumple": cumple, "Documentacion": documentacion, "Controles": controles}


def parece_duda(texto: str) -> bool:
    lowered = texto.strip().lower()
    return lowered.endswith("?") or any(word in lowered for word in PALABRAS_DUDA)


async def explicar(criterio: str, pregunta: str, mensaje: str) -> str:
    respaldo = (
        "Esta pregunta busca saber si la empresa cumple ese punto y cómo puede demostrarlo. "
        "Cuéntame con tus palabras qué hace actualmente la empresa al respecto."
    )
    try:
        return await pedir(
            "Eres un auditor de la Ley 1581 de 2012 que ayuda a resolver una duda.\n"
            f'Criterio: "{criterio}".\nPregunta: "{pregunta}".\n'
            f'El auditado escribió: "{mensaje}".\n\n'
            "Resuelve la duda en dos o tres frases sencillas. Al final invítalo a responder la "
            "pregunta original. No clasifiques todavía."
        )
    except Exception as exc:
        logger.warning("Explicación con Ollama no disponible: %s", exc)
        return respaldo


def categoria(porcentaje: float) -> str:
    if porcentaje >= UMBRAL_ALTO:
        return "Cumplimiento Alto"
    if porcentaje >= UMBRAL_MEDIO:
        return "Cumplimiento Medio"
    return "Cumplimiento Bajo"


def calcular(criterios: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    for criterio in criterios:
        nivel = (
            VALOR[criterio["cumple"]]
            + VALOR[criterio["documentacion"]]
            + VALOR[criterio["controles"]]
        ) / 3
        criterio["nivel"] = round(nivel, 4)
        criterio["puntaje"] = round(nivel * criterio["peso"], 4)

    resumen: list[dict[str, Any]] = []
    for dominio in DOMINIOS:
        items = [criterio for criterio in criterios if criterio["dominio"] == dominio]
        if not items:
            continue
        obtenido = sum(item["puntaje"] for item in items)
        maximo = sum(item["peso"] for item in items)
        porcentaje = obtenido / maximo if maximo else 0
        resumen.append(
            {
                "dominio": dominio,
                "puntaje": round(obtenido, 4),
                "maximo": round(maximo, 4),
                "porcentaje": round(porcentaje, 4),
                "categoria": categoria(porcentaje),
            }
        )

    obtenido_global = sum(item["puntaje"] for item in criterios)
    maximo_global = sum(item["peso"] for item in criterios)
    porcentaje_global = obtenido_global / maximo_global if maximo_global else 0
    global_result = {
        "dominio": "TOTAL GENERAL",
        "puntaje": round(obtenido_global, 4),
        "maximo": round(maximo_global, 4),
        "porcentaje": round(porcentaje_global, 4),
        "categoria": categoria(porcentaje_global),
    }
    return resumen, global_result


def exportar_excel(ruta: Path, criterios: list[dict[str, Any]], resumen: list[dict[str, Any]], global_result: dict[str, Any]) -> None:
    """Crea exactamente las hojas ``evaluacion`` y ``resumen`` solicitadas."""
    head_font = Font(name="Arial", bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="305496")
    centered = Alignment(horizontal="center")

    workbook = Workbook()
    evaluacion = workbook.active
    evaluacion.title = "evaluacion"
    evaluation_headers = [
        "Dominio",
        "Criterio",
        "Importancia",
        "Cumple",
        "Documentación",
        "Controles",
        "Ponderación",
        "Nivel",
        "Puntaje",
    ]
    evaluacion.append(evaluation_headers)
    for cell in evaluacion[1]:
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = centered
    for item in criterios:
        evaluacion.append(
            [
                item["dominio"],
                item["criterio"],
                item["importancia"],
                item["cumple"],
                item["documentacion"],
                item["controles"],
                item["peso"],
                item["nivel"],
                item["puntaje"],
            ]
        )
    for row in evaluacion.iter_rows(min_row=2, min_col=4, max_col=9):
        for cell in row:
            cell.alignment = centered
    evaluacion.freeze_panes = "A2"
    evaluacion.auto_filter.ref = evaluacion.dimensions
    for column, width in {"A": 24, "B": 72, "C": 15, "D": 14, "E": 18, "F": 14, "G": 16, "H": 12, "I": 12}.items():
        evaluacion.column_dimensions[column].width = width

    summary = workbook.create_sheet("resumen")
    summary_headers = ["Dominio", "Puntaje", "Máximo", "% Cumplimiento", "Categoría"]
    summary.append(summary_headers)
    for cell in summary[1]:
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = centered
    for item in [*resumen, global_result]:
        summary.append(
            [
                item["dominio"],
                item["puntaje"],
                item["maximo"],
                item["porcentaje"],
                item["categoria"],
            ]
        )
    for row in summary.iter_rows(min_row=2):
        row[3].number_format = "0.0%"
        for cell in row:
            cell.alignment = centered
    if summary.max_row >= 2:
        for cell in summary[summary.max_row]:
            cell.font = Font(name="Arial", bold=True)
    summary.freeze_panes = "A2"
    summary.auto_filter.ref = summary.dimensions
    for column, width in {"A": 26, "B": 14, "C": 14, "D": 20, "E": 24}.items():
        summary.column_dimensions[column].width = width

    workbook.save(ruta)


def exportar_json(
    ruta: Path,
    profile: "AuditProfile",
    criterios: list[dict[str, Any]],
    resumen: list[dict[str, Any]],
    global_result: dict[str, Any],
) -> None:
    data = {
        "empresa": profile.empresa,
        "fecha": dt.datetime.now(dt.timezone.utc).isoformat(timespec="seconds"),
        "perfil": profile.model_dump(),
        "parametros": {
            "valores": VALOR,
            "umbrales": {"alto": UMBRAL_ALTO, "medio": UMBRAL_MEDIO},
            "pesos_importancia": PESO,
        },
        "global": global_result,
        "dominios": [
            {**row, "criterios": [item for item in criterios if item["dominio"] == row["dominio"]]}
            for row in resumen
        ],
    }
    ruta.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


# ---------------------------------------------------------------------------
# Modelos y sesiones
# ---------------------------------------------------------------------------
class User(BaseModel):
    sub: str
    name: str
    email: str | None = None
    picture: str | None = None
    provider: str


class LoginRequest(BaseModel):
    username: str = Field(min_length=1, max_length=120)
    password: str = Field(min_length=1, max_length=500)


class GoogleLoginRequest(BaseModel):
    credential: str = Field(min_length=20)


class AuthResponse(BaseModel):
    token: str
    user: User


class MeResponse(BaseModel):
    user: User


class AuditProfile(BaseModel):
    empresa: str = Field(min_length=1, max_length=200)
    nit: str = Field(default="", max_length=50)
    tamano_empresa: str = Field(min_length=1, max_length=80)
    sector: str = Field(min_length=1, max_length=160)
    datos_sensibles: bool = False
    datos_menores: bool = False
    transferencias_internacionales: bool = False
    canales_digitales: bool = False
    notas: str = Field(default="", max_length=1000)


class AuditStartRequest(BaseModel):
    profile: AuditProfile


class AuditAnswerRequest(BaseModel):
    audit_id: str = Field(min_length=8)
    answer: str = Field(min_length=1, max_length=5000)


class SummaryRow(BaseModel):
    dominio: str
    puntaje: float
    maximo: float
    porcentaje: float
    categoria: str


class AuditResponse(BaseModel):
    audit_id: str
    question: str | None = None
    message: str | None = None
    dominio: str | None = None
    criterio: str | None = None
    current: int
    total: int
    progress: float
    completed: bool = False
    clarification: bool = False
    excel_url: str | None = None
    json_url: str | None = None
    postgres_saved: bool | None = None
    persistence_warning: str | None = None
    resumen: list[SummaryRow] = Field(default_factory=list)
    resultado_global: SummaryRow | None = None


@dataclass
class AuditSession:
    audit_id: str
    user: User
    profile: AuditProfile
    criterios: list[dict[str, Any]]
    question: str
    index: int = 0
    completed: bool = False
    busy: bool = False
    excel_path: Path | None = None
    json_path: Path | None = None
    postgres_saved: bool | None = None
    persistence_warning: str | None = None
    resumen: list[dict[str, Any]] = field(default_factory=list)
    resultado_global: dict[str, Any] | None = None


AUDIT_SESSIONS: dict[str, AuditSession] = {}


# ---------------------------------------------------------------------------
# Autenticación local y Google
# ---------------------------------------------------------------------------
def _token_secret() -> bytes:
    configured = os.getenv("APP_SECRET", "")
    if configured:
        return configured.encode("utf-8")
    material = f"{GOOGLE_CLIENT_ID}|{LOCAL_AUTH_PASSWORD}|datalex-local-session"
    return hashlib.sha256(material.encode("utf-8")).digest()


def _b64encode(value: bytes) -> str:
    return base64.urlsafe_b64encode(value).decode("ascii").rstrip("=")


def _b64decode(value: str) -> bytes:
    return base64.urlsafe_b64decode(value + "=" * (-len(value) % 4))


def issue_token(user: User) -> str:
    payload = {
        **user.model_dump(),
        "exp": int((dt.datetime.now(dt.timezone.utc) + dt.timedelta(hours=SESSION_TTL_HOURS)).timestamp()),
    }
    body = _b64encode(json.dumps(payload, separators=(",", ":")).encode("utf-8"))
    signature = _b64encode(hmac.new(_token_secret(), body.encode("ascii"), hashlib.sha256).digest())
    return f"{body}.{signature}"


def decode_token(token: str) -> User:
    try:
        body, supplied_signature = token.split(".", 1)
        expected_signature = _b64encode(
            hmac.new(_token_secret(), body.encode("ascii"), hashlib.sha256).digest()
        )
        if not hmac.compare_digest(supplied_signature, expected_signature):
            raise ValueError("firma inválida")
        payload = json.loads(_b64decode(body))
        if int(payload.pop("exp")) < int(dt.datetime.now(dt.timezone.utc).timestamp()):
            raise ValueError("sesión vencida")
        return User.model_validate(payload)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="La sesión no es válida o ya venció",
        ) from exc


async def current_user(
    credentials: Annotated[HTTPAuthorizationCredentials | None, Depends(bearer_scheme)],
) -> User:
    if credentials is None or credentials.scheme.lower() != "bearer":
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Debes iniciar sesión")
    return decode_token(credentials.credentials)


# ---------------------------------------------------------------------------
# PostgreSQL: toda falla se convierte en estado, nunca derriba la aplicación
# ---------------------------------------------------------------------------
def db_config() -> dict[str, str]:
    return {
        "host": os.getenv("DB_HOST", os.getenv("HOST", "localhost")),
        "port": os.getenv("DB_PORT", os.getenv("PORT", "5432")),
        "dbname": os.getenv("DB_NAME", "postgres"),
        "user": os.getenv("DB_USER", "postgres"),
        "password": os.getenv("DB_PASSWORD", ""),
        "connect_timeout": os.getenv("DB_CONNECT_TIMEOUT", "3"),
        "options": "-c client_encoding=UTF8",
    }


def db_connect():
    if psycopg is None:
        raise RuntimeError("psycopg no está instalado")
    return psycopg.connect(**db_config())


def ensure_database(connection: Any) -> None:
    with connection.cursor() as cursor:
        cursor.execute("CREATE SCHEMA IF NOT EXISTS hackaton")
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS hackaton.auditoria (
                auditoria_id TEXT PRIMARY KEY,
                usuario TEXT NOT NULL,
                empresa TEXT NOT NULL,
                nit TEXT,
                sector TEXT,
                tamano_empresa TEXT,
                fecha_hora TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS hackaton.evaluacion (
                auditoria_id TEXT NOT NULL,
                usuario TEXT NOT NULL,
                dominio TEXT NOT NULL,
                criterio TEXT NOT NULL,
                importancia TEXT NOT NULL,
                cumple TEXT NOT NULL,
                documentacion TEXT NOT NULL,
                controles TEXT NOT NULL,
                ponderacion DOUBLE PRECISION NOT NULL,
                nivel DOUBLE PRECISION NOT NULL,
                puntaje DOUBLE PRECISION NOT NULL
            )
            """
        )
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS hackaton.resumen (
                auditoria_id TEXT NOT NULL,
                usuario TEXT NOT NULL,
                dominio TEXT NOT NULL,
                puntaje DOUBLE PRECISION NOT NULL,
                maximo DOUBLE PRECISION NOT NULL,
                porcentaje_cumplimiento DOUBLE PRECISION NOT NULL,
                categoria TEXT NOT NULL
            )
            """
        )


def guardar_resultado_postgres(session: AuditSession) -> tuple[bool, str | None]:
    usuario = session.user.email or session.user.name or session.user.sub
    try:
        with db_connect() as connection:
            ensure_database(connection)
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    INSERT INTO hackaton.auditoria (
                        auditoria_id, usuario, empresa, nit, sector, tamano_empresa
                    ) VALUES (%s, %s, %s, %s, %s, %s)
                    ON CONFLICT (auditoria_id) DO UPDATE SET
                        usuario = EXCLUDED.usuario,
                        empresa = EXCLUDED.empresa,
                        nit = EXCLUDED.nit,
                        sector = EXCLUDED.sector,
                        tamano_empresa = EXCLUDED.tamano_empresa
                    """,
                    (
                        session.audit_id,
                        usuario,
                        session.profile.empresa,
                        session.profile.nit,
                        session.profile.sector,
                        session.profile.tamano_empresa,
                    ),
                )
                cursor.execute("DELETE FROM hackaton.evaluacion WHERE auditoria_id = %s", (session.audit_id,))
                cursor.executemany(
                    """
                    INSERT INTO hackaton.evaluacion (
                        auditoria_id, usuario, dominio, criterio, importancia,
                        cumple, documentacion, controles, ponderacion, nivel, puntaje
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    [
                        (
                            session.audit_id,
                            usuario,
                            item["dominio"],
                            item["criterio"],
                            item["importancia"],
                            item["cumple"],
                            item["documentacion"],
                            item["controles"],
                            item["peso"],
                            item["nivel"],
                            item["puntaje"],
                        )
                        for item in session.criterios
                    ],
                )
                cursor.execute("DELETE FROM hackaton.resumen WHERE auditoria_id = %s", (session.audit_id,))
                cursor.executemany(
                    """
                    INSERT INTO hackaton.resumen (
                        auditoria_id, usuario, dominio, puntaje, maximo,
                        porcentaje_cumplimiento, categoria
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """,
                    [
                        (
                            session.audit_id,
                            usuario,
                            item["dominio"],
                            item["puntaje"],
                            item["maximo"],
                            item["porcentaje"],
                            item["categoria"],
                        )
                        for item in [*session.resumen, session.resultado_global]
                    ],
                )
        return True, None
    except Exception as exc:
        logger.warning("La auditoría %s no pudo guardarse en PostgreSQL: %s", session.audit_id, exc)
        return False, "No se pudo guardar en PostgreSQL; el Excel se generó correctamente y la aplicación continúa activa."


def _safe_filename(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode()
    return re.sub(r"[^A-Za-z0-9_-]+", "_", normalized).strip("_") or "auditoria"


def finalizar_auditoria(session: AuditSession) -> None:
    session.resumen, session.resultado_global = calcular(session.criterios)
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    base = f"auditoria_{_safe_filename(session.profile.empresa)}_{stamp}_{session.audit_id[:8]}"
    session.excel_path = AUDIT_DIR / f"{base}.xlsx"
    session.json_path = AUDIT_DIR / f"{base}.json"
    exportar_excel(
        session.excel_path,
        session.criterios,
        session.resumen,
        session.resultado_global,
    )
    exportar_json(
        session.json_path,
        session.profile,
        session.criterios,
        session.resumen,
        session.resultado_global,
    )
    session.postgres_saved, session.persistence_warning = guardar_resultado_postgres(session)
    session.completed = True


def audit_response(
    session: AuditSession,
    *,
    message: str | None = None,
    clarification: bool = False,
) -> AuditResponse:
    total = len(session.criterios)
    current = total if session.completed else min(session.index + 1, total)
    current_item = None if session.completed else session.criterios[session.index]
    return AuditResponse(
        audit_id=session.audit_id,
        question=None if session.completed else session.question,
        message=message,
        dominio=current_item["dominio"] if current_item else None,
        criterio=current_item["criterio"] if current_item else None,
        current=current,
        total=total,
        progress=round((session.index if not session.completed else total) / total, 4) if total else 1.0,
        completed=session.completed,
        clarification=clarification,
        excel_url=f"/audit/{session.audit_id}/excel" if session.excel_path else None,
        json_url=f"/audit/{session.audit_id}/json" if session.json_path else None,
        postgres_saved=session.postgres_saved,
        persistence_warning=session.persistence_warning,
        resumen=[SummaryRow(**item) for item in session.resumen],
        resultado_global=SummaryRow(**session.resultado_global) if session.resultado_global else None,
    )


# ---------------------------------------------------------------------------
# Rutas API
# ---------------------------------------------------------------------------
@app.get("/health")
async def health() -> dict[str, Any]:
    return {
        "status": "ok",
        "model": OLLAMA_MODEL,
        "frontend": "http://localhost:5173",
        "google_configured": bool(GOOGLE_CLIENT_ID),
    }


@app.get("/auth/config")
async def auth_config() -> dict[str, Any]:
    return {
        "google_client_id": GOOGLE_CLIENT_ID,
        "local_login": True,
        "local_credentials_configured": bool(LOCAL_AUTH_USERNAME and LOCAL_AUTH_PASSWORD),
    }


@app.post("/auth/login", response_model=AuthResponse)
async def local_login(payload: LoginRequest) -> AuthResponse:
    username = payload.username.strip()
    if not username:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="El usuario es obligatorio")

    # Con variables LOCAL_AUTH_* se validan credenciales fijas. Sin ellas se habilita
    # el modo local de demostración y basta con dos campos no vacíos.
    if LOCAL_AUTH_USERNAME or LOCAL_AUTH_PASSWORD:
        valid_user = hmac.compare_digest(username, LOCAL_AUTH_USERNAME)
        valid_password = hmac.compare_digest(payload.password, LOCAL_AUTH_PASSWORD)
        if not (valid_user and valid_password):
            raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Usuario o contraseña incorrectos")

    user = User(sub=f"local:{username}", name=username, provider="local")
    return AuthResponse(token=issue_token(user), user=user)


@app.post("/auth/google", response_model=AuthResponse)
async def google_login(payload: GoogleLoginRequest) -> AuthResponse:
    if not GOOGLE_CLIENT_ID:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="GOOGLE_CLIENT_ID no está configurado",
        )
    try:
        info = await asyncio.to_thread(
            id_token.verify_oauth2_token,
            payload.credential,
            google_requests.Request(),
            GOOGLE_CLIENT_ID,
        )
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Token de Google inválido") from exc

    user = User(
        sub=str(info["sub"]),
        email=info.get("email"),
        name=info.get("name") or info.get("email") or "Usuario de Google",
        picture=info.get("picture"),
        provider="google",
    )
    return AuthResponse(token=issue_token(user), user=user)


@app.get("/auth/me", response_model=MeResponse)
async def me(user: Annotated[User, Depends(current_user)]) -> MeResponse:
    return MeResponse(user=user)


@app.post("/audit/start", response_model=AuditResponse)
async def audit_start(
    payload: AuditStartRequest,
    user: Annotated[User, Depends(current_user)],
) -> AuditResponse:
    profile_text = perfil_a_texto(payload.profile)
    semilla = cargar_semilla()

    # Dos llamadas concurrentes mantienen un inicio razonable sin saturar Ollama local.
    semaphore = asyncio.Semaphore(2)

    async def generate(dominio: str) -> list[dict[str, Any]]:
        async with semaphore:
            return await generar_criterios(profile_text, dominio, semilla.get(dominio, []))

    groups = await asyncio.gather(*(generate(dominio) for dominio in DOMINIOS))
    criterios = [item for group in groups for item in group]
    if not criterios:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="No fue posible construir criterios para la auditoría",
        )

    first_question = await redactar_pregunta(criterios[0]["criterio"])
    audit_id = secrets.token_urlsafe(18)
    session = AuditSession(
        audit_id=audit_id,
        user=user,
        profile=payload.profile,
        criterios=criterios,
        question=first_question,
    )
    AUDIT_SESSIONS[audit_id] = session
    return audit_response(session, message="La auditoría fue creada. Responde con tus propias palabras.")


@app.post("/audit/answer", response_model=AuditResponse)
async def audit_answer(
    payload: AuditAnswerRequest,
    user: Annotated[User, Depends(current_user)],
) -> AuditResponse:
    session = AUDIT_SESSIONS.get(payload.audit_id)
    if session is None or session.user.sub != user.sub:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Auditoría no encontrada")
    if session.completed:
        return audit_response(session, message="La auditoría ya está completa.")
    if session.busy:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="La respuesta anterior aún se está procesando")

    session.busy = True
    try:
        current = session.criterios[session.index]
        if parece_duda(payload.answer):
            explanation = await explicar(current["criterio"], session.question, payload.answer)
            return audit_response(session, message=explanation, clarification=True)

        result = await clasificar(current["criterio"], payload.answer)
        if result is None:
            result = clasificar_respaldo(payload.answer)
        current["respuesta"] = payload.answer
        current["cumple"] = result["Cumple"]
        current["documentacion"] = result["Documentacion"]
        current["controles"] = result["Controles"]

        session.index += 1
        if session.index >= len(session.criterios):
            await asyncio.to_thread(finalizar_auditoria, session)
            return audit_response(
                session,
                message="Evaluación finalizada. El archivo Excel ya está listo para descargar.",
            )

        session.question = await redactar_pregunta(session.criterios[session.index]["criterio"])
        return audit_response(session, message="Respuesta registrada.")
    finally:
        session.busy = False


def _owned_session(audit_id: str, user: User) -> AuditSession:
    session = AUDIT_SESSIONS.get(audit_id)
    if session is None or session.user.sub != user.sub:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Auditoría no encontrada")
    return session


@app.get("/audit/{audit_id}/excel")
async def download_excel(
    audit_id: str,
    user: Annotated[User, Depends(current_user)],
) -> FileResponse:
    session = _owned_session(audit_id, user)
    if session.excel_path is None or not session.excel_path.exists():
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="El Excel aún no está disponible")
    return FileResponse(
        session.excel_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=session.excel_path.name,
    )


@app.get("/audit/{audit_id}/json")
async def download_json(
    audit_id: str,
    user: Annotated[User, Depends(current_user)],
) -> FileResponse:
    session = _owned_session(audit_id, user)
    if session.json_path is None or not session.json_path.exists():
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="El JSON aún no está disponible")
    return FileResponse(session.json_path, media_type="application/json", filename=session.json_path.name)


# El frontend estático se sirve desde el mismo proceso y puerto. Debe ir al final
# para que las rutas API anteriores tengan prioridad.
if FRONTEND_DIR.exists():
    app.mount("/", StaticFiles(directory=FRONTEND_DIR, html=True), name="frontend")


if __name__ == "__main__":
    import uvicorn

    uvicorn.run("main:app", host="127.0.0.1", port=8000, reload=True)
