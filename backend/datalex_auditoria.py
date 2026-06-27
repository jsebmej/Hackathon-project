#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
DATALEX-1581 - Auditoría adaptativa (generación + clasificación + puntuación)
-----------------------------------------------------------------------------
Flujo (lo controla este script; el modelo solo ejecuta micro-tareas):
  1. Toma el perfil de la empresa.
  2. El modelo GENERA criterios a la medida por dominio y les asigna importancia.
  3. Para cada criterio el modelo redacta una pregunta distinta; tú respondes libre.
  4. El modelo CLASIFICA la respuesta en Cumple / Documentación / Controles (Sí/No/Parcial).
  5. El script calcula puntajes y categorías (determinístico, no lo hace el modelo).
  6. Exporta:  auditoria_<empresa>_<fecha>.xlsx   (motor con fórmulas vivas)
               auditoria_<empresa>_<fecha>.json  (para alimentar tu front)

USO:
    python datalex_auditoria.py "Nombre Empresa"

Requisitos:  pip install requests openpyxl   y Ollama corriendo en localhost:11434
"""

import sys
import re
import json
import unicodedata
import datetime
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ----------------------------------------------------------------------
# CONFIGURACIÓN
# ----------------------------------------------------------------------
OLLAMA_URL = "http://localhost:11434/api/chat"
MODELO = "datalex1581"          # usa un modelo 7B+ para mejores criterios

DOMINIOS = ["Consentimiento", "Finalidad", "Seguridad",
            "Derechos del titular", "Gestión documental"]

# Cuántos criterios pedir por dominio (el modelo se ajusta dentro del rango).
CRIT_MIN, CRIT_MAX = 4, 6

# Importancia -> peso (ponderación). El modelo asigna la importancia.
PESO = {"Alta": 3, "Media": 2, "Baja": 1}

# Valor numérico de cada respuesta.
VALOR = {"Sí": 1.0, "Parcial": 0.5, "No": 0.0}

# Cortes de categoría (sobre el % de cumplimiento, 0–1).
UMBRAL_ALTO, UMBRAL_MEDIO = 0.80, 0.50

PREGUNTA_CON_MODELO = True       # False = usa el texto del criterio como pregunta

# Tope de criterios por dominio tras la generación (evita listas enormes).
TOPE_POR_DOMINIO = 10

# Archivo opcional para sobrescribir la semilla embebida. Si existe en esta
# carpeta, se leen de ahí las columnas Dominio y Criterio.
ARCHIVO_SEMILLA = "20260626_-_Criterios_de_aceptacion.xlsx"

# SEMILLA: tus criterios base por dominio (anclan la generación a tu matriz).
# Edita esta lista para curar tu base; el modelo parte de aquí y la adapta.
SEMILLA = {
    "Consentimiento": [
        "La empresa solicita autorización previa del titular",
        "La autorización es informada y clara",
        "Se conserva evidencia del consentimiento",
        "Permite revocar el consentimiento",
        "El consentimiento es verificable digital o físicamente",
        "No trata datos sin autorización válida",
        "Existen formatos de autorización estandarizados",
        "Se valida el consentimiento antes del tratamiento",
        "Se actualizan consentimientos cuando cambia la finalidad",
        "Se informa al titular sobre el uso de datos",
    ],
    "Finalidad": [
        "Los datos tienen una finalidad claramente definida",
        "La finalidad se informa al titular antes del tratamiento",
        "No se usan datos para fines distintos a los autorizados",
        "La finalidad está documentada en políticas internas",
        "Se revisa periódicamente la finalidad del tratamiento",
        "Los sistemas validan el uso según la finalidad",
        "Se eliminan datos cuando se cumple la finalidad",
        "Los empleados conocen la finalidad del tratamiento",
        "Se audita el cumplimiento de la finalidad",
        "Se actualizan finalidades cuando cambian procesos",
    ],
    "Seguridad": [
        "Existen controles de acceso a bases de datos",
        "Se usa cifrado de información sensible",
        "Hay copias de seguridad periódicas",
        "Existen políticas de seguridad de la información",
        "Se monitorean accesos no autorizados",
        "Se gestionan incidentes de seguridad",
        "Se restringe acceso según roles",
        "Se protegen credenciales de acceso",
        "Se realizan pruebas de seguridad periódicas",
        "Los sistemas están protegidos contra ataques externos",
    ],
    "Derechos del titular": [
        "La empresa responde solicitudes de consulta",
        "La empresa responde solicitudes de reclamo",
        "Permite actualización de datos personales",
        "Permite eliminación de datos",
        "Existen canales de atención al titular",
        "Se respetan tiempos legales de respuesta",
        "Se registra cada solicitud del titular",
        "Se verifica identidad del solicitante",
        "Se informa al titular sobre el uso de sus datos",
        "Existe procedimiento documentado de derechos",
    ],
    "Gestión documental": [
        "Existe política de tratamiento de datos",
        "La política está publicada y accesible",
        "Se actualiza la política periódicamente",
        "Se documentan bases de datos",
        "Existe inventario de datos personales",
        "Se documentan responsables del tratamiento",
        "Se documentan encargados externos",
        "Se realizan auditorías internas",
        "Se conservan evidencias de cumplimiento",
        "Existe control de versiones de documentos",
    ],
}


def cargar_semilla():
    """Si existe ARCHIVO_SEMILLA, lee la base desde ahí; si no, usa la embebida."""
    import os
    if not os.path.exists(ARCHIVO_SEMILLA):
        return SEMILLA
    try:
        from openpyxl import load_workbook
        ws = load_workbook(ARCHIVO_SEMILLA, read_only=True, data_only=True).active
        base = {}
        for fila in ws.iter_rows(min_row=2, values_only=True):
            dom, crit = (fila[0] if fila else None), (fila[1] if len(fila) > 1 else None)
            if dom and crit:
                base.setdefault(str(dom).strip(), []).append(str(crit).strip())
        return base or SEMILLA
    except Exception:
        return SEMILLA


# ----------------------------------------------------------------------
# UTILIDADES DE TEXTO / PARSEO
# ----------------------------------------------------------------------
def norm(s):
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]", "", s.lower())


def normaliza_valor(v):
    n = norm(v)
    if n == "si":
        return "Sí"
    if n.startswith("par"):
        return "Parcial"
    if n == "no":
        return "No"
    return None


def normaliza_importancia(v):
    n = norm(v)
    if n.startswith("alt"):
        return "Alta"
    if n.startswith("baj"):
        return "Baja"
    return "Media"


def _limpia(texto):
    return re.sub(r"```[a-zA-Z]*", "", texto).replace("```", "")


def extraer_array(texto):
    t = _limpia(texto)
    i, f = t.find("["), t.rfind("]")
    if i == -1 or f == -1:
        return None
    frag = re.sub(r",(\s*[\]}])", r"\1", t[i:f + 1])
    try:
        return json.loads(frag)
    except json.JSONDecodeError:
        return None


def extraer_objeto(texto):
    t = _limpia(texto)
    i, f = t.find("{"), t.rfind("}")
    if i == -1 or f == -1:
        return None
    frag = re.sub(r",(\s*[}\]])", r"\1", t[i:f + 1])
    try:
        return json.loads(frag)
    except json.JSONDecodeError:
        return None


# ----------------------------------------------------------------------
# OLLAMA (una llamada por micro-tarea, sin historial)
# ----------------------------------------------------------------------
def pedir(instruccion):
    payload = {"model": MODELO,
               "messages": [{"role": "user", "content": instruccion}],
               "stream": False}
    r = requests.post(OLLAMA_URL, json=payload, timeout=300)
    if r.status_code != 200:
        # Ollama devuelve el motivo en el cuerpo (p. ej. "model not found").
        raise RuntimeError(f"Ollama {r.status_code}: {r.text[:300]}")
    return r.json().get("message", {}).get("content", "").strip()


def generar_criterios(perfil, dominio, semilla_dom):
    base = "\n".join(f"- {c}" for c in semilla_dom)
    instr = (
        f"Perfil de la empresa:\n{perfil}\n\n"
        f"Dominio: \"{dominio}\" (Ley 1581 de 2012).\n"
        f"Criterios base de auditoría para este dominio:\n{base}\n\n"
        "Tu tarea: SELECCIONA de la lista base los criterios que apliquen a ESTA empresa "
        "según su perfil, reescríbelos si hace falta para hacerlos específicos, y AGREGA "
        "criterios nuevos propios de esta empresa por su sector, tipos de datos y riesgos. "
        "No inventes artículos ni sanciones. Prioriza los más relevantes "
        f"(alrededor de {CRIT_MIN} a {CRIT_MAX}). Asigna a cada criterio una importancia.\n"
        'Responde SOLO este JSON: [{"criterio":"...","importancia":"Alta|Media|Baja"}]'
    )
    try:
        arr = extraer_array(pedir(instr)) or []
    except Exception:
        arr = []
    salida = []
    for it in arr:
        if isinstance(it, dict):
            crit = next((v for k, v in it.items() if norm(k) == "criterio"), None)
            imp = next((v for k, v in it.items() if norm(k) == "importancia"), None)
            if crit and str(crit).strip():
                salida.append({"criterio": str(crit).strip(),
                               "importancia": normaliza_importancia(imp)})
    if not salida:
        # Respaldo: si el modelo falla, usa tus criterios base tal cual.
        salida = [{"criterio": c, "importancia": "Media"} for c in semilla_dom]
    return salida[:TOPE_POR_DOMINIO]


def redactar_pregunta(criterio):
    if not PREGUNTA_CON_MODELO:
        return f"Respecto a «{criterio}», ¿cuál es la situación en la empresa?"
    try:
        q = pedir(
            "Escribe UNA sola pregunta clara y natural, en español, para verificar "
            f"en una empresa este criterio: \"{criterio}\". Devuelve solo la pregunta."
        )
        return q.split("\n")[0].strip() or f"Sobre «{criterio}», ¿qué hace la empresa?"
    except Exception:
        return f"Sobre «{criterio}», ¿qué hace la empresa?"


def clasificar(criterio, respuesta):
    instr = (
        f"Criterio: \"{criterio}\".\nRespuesta del auditado: \"{respuesta}\".\n\n"
        "Clasifica en tres dimensiones y responde SOLO este JSON:\n"
        '{"Cumple":"Sí|No|Parcial","Documentacion":"Sí|No|Parcial","Controles":"Sí|No|Parcial"}\n'
        "Cumple = la empresa cumple el criterio. Documentacion = hay soporte documental. "
        "Controles = hay controles que lo aseguren. Solo Sí, No o Parcial. "
        "Si algo no se menciona, no lo asumas como Sí."
    )
    try:
        obj = extraer_objeto(pedir(instr))
    except Exception:
        obj = None
    if not obj:
        return None
    res = {}
    for clave, dest in (("cumple", "Cumple"),
                        ("documentacion", "Documentacion"),
                        ("controles", "Controles")):
        valor = next((v for k, v in obj.items() if norm(k) == clave), None)
        res[dest] = normaliza_valor(valor)
    return res if None not in res.values() else None


def clasificar_manual(criterio):
    print(f"   (Clasificación manual de: {criterio})")
    out = {}
    for dim in ("Cumple", "Documentacion", "Controles"):
        while True:
            v = normaliza_valor(input(f"   {dim} (Si/No/Parcial)> "))
            if v:
                out[dim] = v
                break
            print("   Escribe Si, No o Parcial.")
    return out


# Señales de que el auditado tiene una duda o pregunta (en vez de responder).
PALABRAS_DUDA = ("no entiendo", "no se", "no sé", "que es", "qué es", "explica",
                 "explicame", "explícame", "aclara", "ayuda", "no comprendo",
                 "como asi", "cómo así", "a que te refieres", "a qué te refieres",
                 "no me queda claro", "ejemplo", "puedes explicar")


def parece_duda(texto):
    t = texto.strip().lower()
    return t.endswith("?") or any(p in t for p in PALABRAS_DUDA)


def explicar(criterio, pregunta, mensaje):
    """El LLM resuelve la duda del auditado y vuelve a invitar a responder."""
    instr = (
        "Eres un auditor de la Ley 1581 de 2012 que ayuda a un auditado con dudas.\n"
        f"Criterio que se verifica: \"{criterio}\".\n"
        f"Pregunta original: \"{pregunta}\".\n"
        f"El auditado escribió: \"{mensaje}\".\n\n"
        "Resuelve su duda en 2 o 3 frases claras y sencillas, sin tecnicismos. "
        "Al final, vuelve a invitarlo a responder la pregunta. No clasifiques todavía."
    )
    try:
        return pedir(instr) or ("Esa pregunta busca saber si la empresa cumple ese punto. "
                                "Cuéntame con tus palabras qué hace la empresa al respecto.")
    except Exception:
        return ("Esa pregunta busca saber si la empresa cumple ese punto. "
                "Cuéntame con tus palabras qué hace la empresa al respecto.")


def conversar_criterio(criterio, pregunta):
    """Mini-diálogo por criterio: aclara dudas con el LLM, clasifica al haber respuesta."""
    MAX = 6
    for intento in range(1, MAX + 1):
        resp = input("Tu respuesta> ").strip()
        low = resp.lower()
        if low == "/salir":
            return {"_salir": True}
        if low == "/saltar":
            return {"Cumple": "No", "Documentacion": "No", "Controles": "No"}
        if low == "/manual":
            return clasificar_manual(criterio)
        if not resp:
            continue
        # Si es una duda/pregunta, el modelo explica y volvemos a preguntar.
        if parece_duda(resp) and intento < MAX:
            print(f"\nDATALEX> {explicar(criterio, pregunta, resp)}")
            continue
        # Si parece respuesta, intentamos clasificar.
        cls = clasificar(criterio, resp)
        if cls:
            return cls
        # No se pudo interpretar: si quedan intentos, pedimos precisar; si no, manual.
        if intento < MAX:
            print(f"\nDATALEX> {explicar(criterio, pregunta, resp)}")
            continue
        print("   No logré interpretar la respuesta; clasifícalo tú "
              "(o escribe /manual antes):")
        return clasificar_manual(criterio)
    return clasificar_manual(criterio)


# ----------------------------------------------------------------------
# PUNTUACIÓN (determinística, en Python)
# ----------------------------------------------------------------------
def categoria(pct):
    if pct >= UMBRAL_ALTO:
        return "Cumplimiento Alto"
    if pct >= UMBRAL_MEDIO:
        return "Cumplimiento Medio"
    return "Cumplimiento Bajo"


def calcular(criterios):
    for c in criterios:
        nivel = (VALOR[c["cumple"]] + VALOR[c["documentacion"]] + VALOR[c["controles"]]) / 3
        c["nivel"] = round(nivel, 4)
        c["puntaje"] = round(nivel * c["peso"], 4)
    resumen = []
    for dom in DOMINIOS:
        items = [c for c in criterios if c["dominio"] == dom]
        if not items:
            continue
        obt = sum(c["puntaje"] for c in items)
        maxi = sum(c["peso"] for c in items)
        pct = obt / maxi if maxi else 0
        resumen.append({"dominio": dom, "puntaje": round(obt, 4),
                        "maximo": round(maxi, 4), "porcentaje": round(pct, 4),
                        "categoria": categoria(pct)})
    g_obt = sum(c["puntaje"] for c in criterios)
    g_max = sum(c["peso"] for c in criterios)
    g_pct = g_obt / g_max if g_max else 0
    global_ = {"puntaje": round(g_obt, 4), "maximo": round(g_max, 4),
               "porcentaje": round(g_pct, 4), "categoria": categoria(g_pct)}
    return resumen, global_


# ----------------------------------------------------------------------
# EXPORTACIÓN
# ----------------------------------------------------------------------
def exportar_json(ruta, empresa, perfil, criterios, resumen, global_):
    data = {
        "empresa": empresa,
        "fecha": datetime.datetime.now().isoformat(timespec="seconds"),
        "perfil": perfil,
        "parametros": {"valores": VALOR,
                       "umbrales": {"alto": UMBRAL_ALTO, "medio": UMBRAL_MEDIO},
                       "pesos_importancia": PESO},
        "global": global_,
        "dominios": [{**d, "criterios": [c for c in criterios if c["dominio"] == d["dominio"]]}
                     for d in resumen],
    }
    with open(ruta, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def exportar_excel(ruta, criterios):
    HEAD = Font(name="Arial", bold=True, color="FFFFFF")
    FILL = PatternFill("solid", fgColor="305496")
    AZUL = Font(name="Arial", color="0000FF")
    NEG = Font(name="Arial", color="000000")
    BOLD = Font(name="Arial", bold=True)
    CEN = Alignment(horizontal="center")

    wb = Workbook()
    ev = wb.active
    ev.title = "Evaluacion"
    cols = ["Dominio", "Criterio", "Importancia", "Cumple", "Documentación",
            "Controles", "Ponderación", "Nivel", "Puntaje"]
    for j, h in enumerate(cols, 1):
        c = ev.cell(1, j, h); c.font = HEAD; c.fill = FILL; c.alignment = CEN

    for i, c in enumerate(criterios):
        r = i + 2
        ev.cell(r, 1, c["dominio"]).font = NEG
        ev.cell(r, 2, c["criterio"]).font = NEG
        ev.cell(r, 3, c["importancia"]).font = NEG
        for j, k in ((4, "cumple"), (5, "documentacion"), (6, "controles")):
            cc = ev.cell(r, j, c[k]); cc.font = AZUL; cc.alignment = CEN
        ev.cell(r, 7, c["peso"]).font = AZUL
        # Nivel = promedio de las 3 dimensiones (Sí=1, Parcial=0.5, No=0)
        conv = ('IF({0}="Sí",1,IF({0}="Parcial",0.5,0))')
        ev.cell(r, 8).value = f"=({conv.format(f'D{r}')}+{conv.format(f'E{r}')}+{conv.format(f'F{r}')})/3"
        ev.cell(r, 8).font = NEG; ev.cell(r, 8).number_format = "0.00"; ev.cell(r, 8).alignment = CEN
        ev.cell(r, 9).value = f"=H{r}*G{r}"
        ev.cell(r, 9).font = NEG; ev.cell(r, 9).number_format = "0.000"; ev.cell(r, 9).alignment = CEN

    ult = len(criterios) + 1
    ev.column_dimensions["B"].width = 60
    ev.column_dimensions["A"].width = 20
    ev.freeze_panes = "A2"

    rs = wb.create_sheet("Resumen")
    for j, h in enumerate(["Dominio", "Puntaje", "Máximo", "% Cumplimiento", "Categoría"], 1):
        c = rs.cell(1, j, h); c.font = HEAD; c.fill = FILL; c.alignment = CEN
    dom_rng = f"Evaluacion!$A$2:$A${ult}"
    pun_rng = f"Evaluacion!$I$2:$I${ult}"
    pon_rng = f"Evaluacion!$G$2:$G${ult}"
    cat = ('=IF(D{0}>={a},"Cumplimiento Alto",IF(D{0}>={m},"Cumplimiento Medio","Cumplimiento Bajo"))'
           .replace("{a}", str(UMBRAL_ALTO)).replace("{m}", str(UMBRAL_MEDIO)))
    doms = list(dict.fromkeys(c["dominio"] for c in criterios))
    for i, dom in enumerate(doms):
        r = i + 2
        rs.cell(r, 1, dom).font = NEG
        rs.cell(r, 2).value = f"=SUMIF({dom_rng},A{r},{pun_rng})"
        rs.cell(r, 3).value = f"=SUMIF({dom_rng},A{r},{pon_rng})"
        rs.cell(r, 4).value = f"=IF(C{r}=0,0,B{r}/C{r})"
        rs.cell(r, 5).value = cat.format(r)
        rs.cell(r, 4).number_format = "0.0%"
        for j in range(1, 6): rs.cell(r, j).font = NEG
    g = len(doms) + 2
    rs.cell(g, 1, "TOTAL GENERAL").font = BOLD
    rs.cell(g, 2).value = f"=SUM(B2:B{g-1})"
    rs.cell(g, 3).value = f"=SUM(C2:C{g-1})"
    rs.cell(g, 4).value = f"=IF(C{g}=0,0,B{g}/C{g})"
    rs.cell(g, 5).value = cat.format(g)
    rs.cell(g, 4).number_format = "0.0%"
    for j in range(1, 6): rs.cell(g, j).font = BOLD
    rs.column_dimensions["A"].width = 22
    for L in "BCDE": rs.column_dimensions[L].width = 16
    wb.save(ruta)


# ----------------------------------------------------------------------
# INTAKE DE PERFIL
# ----------------------------------------------------------------------
def si_no(p):
    return "Sí" if input(p).strip().lower().startswith("s") else "No"


def tomar_perfil():
    print("--- Perfil de la empresa (define los criterios a la medida) ---")
    sector = input("Sector / actividad económica: ").strip()
    tam = input("Tamaño (micro/pequeña/mediana/grande): ").strip()
    sens = si_no("¿Maneja datos sensibles (salud, biométricos, etc.)? (s/n): ")
    menores = si_no("¿Trata datos de menores de edad? (s/n): ")
    inter = si_no("¿Hace transferencias o almacenamiento internacional? (s/n): ")
    digital = si_no("¿Tiene canales digitales / app / e-commerce? (s/n): ")
    extra = input("Notas adicionales (opcional): ").strip()
    return (f"Sector: {sector}. Tamaño: {tam}. Datos sensibles: {sens}. "
            f"Datos de menores: {menores}. Transferencias internacionales: {inter}. "
            f"Canales digitales: {digital}. Notas: {extra or 'ninguna'}.")


# ----------------------------------------------------------------------
# PRINCIPAL
# ----------------------------------------------------------------------
def main():
    empresa = sys.argv[1] if len(sys.argv) > 1 else "auditoria"
    sello = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    base = f"auditoria_{empresa.replace(' ', '_')}_{sello}"

    try:
        requests.get("http://localhost:11434", timeout=10)
    except Exception:
        print("[ERROR] No hay conexión con Ollama en localhost:11434. Ábrelo y reintenta.")
        return

    # Autodiagnóstico: si el modelo no responde, detente aquí (no sigas en manual).
    print(f"Probando el modelo «{MODELO}»...")
    try:
        prueba = pedir("Responde únicamente la palabra: ok")
    except Exception as e:
        print(f"[ERROR] El modelo no respondió: {e}")
        print(f"  Revisa:   ollama list        (¿aparece «{MODELO}»?)")
        print(f"  Prueba:   ollama run {MODELO} \"hola\"")
        print(f"  Si falta: ollama create {MODELO} -f DataLex_ajustado.Modelfile")
        return
    if not prueba:
        print("[ERROR] El modelo respondió VACÍO.")
        print("  Suele pasar si el 'stop' del Modelfile no corresponde al modelo base.")
        print("  Si cambiaste el FROM a otro modelo (qwen, etc.), quita o ajusta la línea PARAMETER stop.")
        return
    print(f"Modelo OK -> {prueba[:60]}\n")

    perfil = tomar_perfil()
    semilla = cargar_semilla()
    print(f"\nGenerando criterios a la medida para «{empresa}» (anclados a tu matriz)...\n")

    criterios = []
    for dom in DOMINIOS:
        try:
            gen = generar_criterios(perfil, dom, semilla.get(dom, []))
        except Exception as e:
            print(f"[aviso] error generando criterios de {dom}: {e}")
            gen = [{"criterio": c, "importancia": "Media"} for c in semilla.get(dom, [])]
        if not gen:
            print(f"[aviso] no hay criterios para {dom}; se omite.")
            continue
        print(f"== {dom}: {len(gen)} criterios generados ==")
        for g in gen:
            criterios.append({"dominio": dom, "criterio": g["criterio"],
                              "importancia": g["importancia"],
                              "peso": PESO[g["importancia"]]})

    if not criterios:
        print("No se generó ningún criterio. Revisa que el modelo responda en JSON.")
        return

    print(f"\n--- Entrevista ({len(criterios)} criterios) ---")
    print("Puedes pedir aclaraciones; el modelo te explica. Comandos: /saltar  /manual  /salir\n")
    dom_actual = None
    for c in criterios:
        if c["dominio"] != dom_actual:
            dom_actual = c["dominio"]
            print(f"\n===== {dom_actual} =====")
        pregunta = redactar_pregunta(c["criterio"])
        print(f"\nDATALEX> {pregunta}")
        cls = conversar_criterio(c["criterio"], pregunta)
        if cls.get("_salir"):
            print("Terminado antes de tiempo; se evalúa lo respondido.")
            for r in criterios:
                r.setdefault("cumple", "No"); r.setdefault("documentacion", "No"); r.setdefault("controles", "No")
            break
        c["cumple"], c["documentacion"], c["controles"] = cls["Cumple"], cls["Documentacion"], cls["Controles"]
        print(f"   Registrado -> Cumple: {c['cumple']} | Doc: {c['documentacion']} | Controles: {c['controles']}")

    # asegura que todos tengan valores (por si /salir)
    for c in criterios:
        c.setdefault("cumple", "No"); c.setdefault("documentacion", "No"); c.setdefault("controles", "No")

    resumen, global_ = calcular(criterios)
    exportar_excel(base + ".xlsx", criterios)
    exportar_json(base + ".json", empresa, perfil, criterios, resumen, global_)

    print("\n===== RESULTADO =====")
    for d in resumen:
        print(f"  {d['dominio']:<22} {d['porcentaje']*100:5.1f}%  {d['categoria']}")
    print(f"  {'GLOBAL':<22} {global_['porcentaje']*100:5.1f}%  {global_['categoria']}")
    print(f"\nArchivos generados:\n  {base}.xlsx  (motor de puntuación)\n  {base}.json  (para tu front)")


if __name__ == "__main__":
    main()
